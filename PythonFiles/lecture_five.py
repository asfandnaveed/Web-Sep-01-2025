# Nested Functions
# File handling
# Exception Handling

# Coursera Assignment (File Handling)


# 1. Nested Functions
# non-local scopes

# def outer():
#     x = 'outer'
#     def inner(): #nested function
#         nonlocal x
#         x = 'inner'
#         print(x)

#     # outer block
#     inner()
#     print(x)

# outer()

# 2. File Handling
# working with txt files

# writing to a file
file = open('users.txt', 'w')
file.write('User 1: Ahmed')
file.close()

# appending to a file
file = open('users.txt', 'a')
file.write('User 2: Faizan')
file.close()

# function with file handling

def fileWrite(filename):
    with open(filename, 'w') as file:
        file.write('A line written by Python')

fileWrite('sample.txt')

# read content from a file and store it in a list

with open('data.txt') as file:
    file_content = file.read()
    content_list = file_content.split('\n')
    print(content_list)

# working with excel files
from openpyxl import load_workbook

workbook = load_workbook('students.xlsx')
active_sheet = workbook.active

for row in active_sheet.iter_rows(values_only = True):
    
    print(row)




        